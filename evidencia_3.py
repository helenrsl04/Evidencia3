contador = {   "Clientes": 0,                            
               "Salas": 0,                               
               "Programacion": 30
            }

cliente =  {   "CliCod": 0,
               "CliNam": 'Por Definir'              
            }


sala =     {   "sala 0": 'Sala Cero'       
           }

salaCapacidad =     {   "sala 0": '0'       
           }


agenda =   {   "0": '0'       
           }

#manejo de SQL Lite 3 
#20221029
import sys
import sqlite3
from sqlite3 import Error
#manejo de SQL Lite 3



import openpyxl                                                                 
import csv

def f_EncontroCita(cita,agenda):                                                
   #### Validacion que no se repita Fecha-Sala-Turno 15-Oct-22                  
                  v_validacion = False                                          
                                                                                
                  for clave in agenda:                                          
                      if (len(clave) >= 10):                                    
                          if clave[0:10] == cita:                               
                             v_validacion = True                                

                  return v_validacion                                           
   #### Validacion que no se repita Fecha-Sala-Turno 15-Oct-22


def f_ValidaCadenaVacia(cadena):                                                
  if (len(cadena) == 0):
    return True                                                                 
  else:                                                                        
    return False                                                                


#manejo de SQL Lite 3 
#20221030
def p_CreacionTablas():
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_contador (ConCod TEXT    , ConUlt INTEGER );")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_cliente  (CliCod INTEGER , CliNom TEXT );")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_sala     (SalCod INTEGER , SalNom TEXT ,SalCap TEXT);")          
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_agenda   (AgeCod TEXT ,    AgeNom TEXT);")                    

          mi_cursor.execute("DELETE FROM t_contador;")
          mi_cursor.execute("DELETE FROM t_cliente ;")
          mi_cursor.execute("DELETE FROM t_sala    ;")          
          mi_cursor.execute("DELETE FROM t_agenda  ;")                    


          #mi_cursor.execute("CREATE TABLE IF NOT EXISTS proyecto (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, responsable INTEGER NOT NULL, FOREIGN KEY(responsable) REFERENCES responsable(clave));")
          print("Tablas creadas exitosamente")
  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
  finally:
      conn.close()      

def f_SiguienteValorContador(valor_clave):
  #Lee todos los datos de una tabla que cumplan una condición
  #valor_clave = int(input("Clave del proyecto a consultar: "))

  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          valores = {"ConCod":valor_clave}
          mi_cursor.execute("SELECT * FROM t_contador WHERE ConCod = :ConCod", valores)
          registro = mi_cursor.fetchall()

          if registro:
              for ConCod, ConUlt in registro:
                  #print(f"{clave}\t{nombre}")
                  #Incrementa el contador
                  return ConUlt
          else:
              #print(f"No se encontró un proyecto asociado con la clave {valor_clave}")
              #Inserta el registro con valor 1 y devuelve 1
              return 0

  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")  


def f_ValidaValorEnTabla(valor_clave):
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          valores = {"ConCod":valor_clave}
          mi_cursor.execute("SELECT * FROM t_contador WHERE ConCod = :ConCod", valores)
          registro = mi_cursor.fetchall()

          if registro:
              for ConCod, ConUlt in registro:
                  #print(f"{clave}\t{nombre}")
                  #Incrementa el contador
                  return ConUlt
          else:
              #print(f"No se encontró un proyecto asociado con la clave {valor_clave}")
              #Inserta el registro con valor 1 y devuelve 1
              return 0

  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")  



def p_EjecutaComandoSQL(SentenciaSql,Parametros):
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute(SentenciaSql, Parametros)
          #registros = mi_cursor.fetchall()
          #if registros:
            #return len(registros)
          #else:
            #return 0
          print("SQL Ejecutado _Exitosamente")
  except Error as e:
      print("Mensaje Error")
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
  finally:
      conn.close()  
#manejo de SQL Lite 3 


#  ************* MENU PRINCIPAL ****************
p_CreacionTablas()

opcion = '0'                                                                    
print("") #Espacio
while not(opcion=='5'):                                                         
    print("      **** MENU PRINCIPAL ****  ")
    print(' [1]. Reservaciones ')
    print(' [2]. Reportes')
    print(' [3]. Registrar una sala')
    print(' [4]. Registrar un cliente')
    print(' [5]. Salir')
    print("")

    opcion=input('  --- ¿Cuál opcion?: ')
#  ************* MENU PRINCIPAL ****************

#  ************* [1]. RESERVACIONES ****************
#  ************* SUBMENU 1. RESERVACIONES ****************
    if (opcion=='1'):                                                           
        opcion = '0'                                                            

        while not(opcion=='9'):
            print('          *** SUBMENU DE RESERVACIONES***             ')
            print(' [1]. Registrar una nueva reservación ')                     
            print(' [2]. Modificiar descripción de una reservación ') 
            print(' [3]. Consultar disponibilidad de salas para una fecha ')
            print(' [4]. Eliminar una reservacion') #Evidencia 3 nuevo
            print(' [5]. Salir')                                                
            print("") #espacio
#  ************* SUBMENU 1. RESERVACIONES ****************
            opcion=input('  --- ¿Cuál opcion?: ')


#  ************* SUBMENU 1. [1]. Registrar una nueva reservación ****************              
            if (opcion=='1'):
                print('Registrar una nueva reservación')


                ########## Validacion SALA #########
                while True:                                                     
                  try:                                                          
                    iSala = int(input("SALA ?: "))                              
                  except ValueError:                                            
                    print("") #Espacio
                    print("Debes escribir un número.")
                    continue                                                    

                  if (iSala < 0):                                               
                    print("Debes escribir un número positivo.")          
                    continue                                                    
                  else:                                                         
                    #break
                    if (iSala in sala):                                         
                      validar = sala.get(iSala)                                 
                      print(validar)
                      break
                    else:
                      print("No Existe sala " , iSala)
                      continue
                ########## Validacion SALA #########


                ########## Validacion CLIENTE #########
                while True:
                  try:
                    iCliente = int(input("CLIENTE ?: "))
                  except ValueError:
                    print("Debes escribir un número.")
                    continue

                  if (iCliente < 0):
                    print("Debes escribir un número positivo.")
                    continue
                  else:
                    #break
                    if (iCliente in cliente):
                      validar = cliente.get(iCliente)
                      print(validar)
                      print("") #Espacio
                      break
                    else:
                      print("No Existe cliente " , iCliente)
                      continue
                ########## Validacion CLIENTE #########


                ########## Validacion TURNO #########
                while True:
                    iTurno = input("TURNO [M]añana [T]arde  [N]oche ?: ")

                    iTurno = iTurno.upper()

                    if (iTurno != "M" and iTurno != "T" and iTurno != "N"):
                      print("Turno no valido debe ser [M] [T] [N]")
                      continue
                    else:
                      break
                ########## Validacion TURNO #########
                      

                ########## Validacion FECHA #########
                while True:
                    from datetime import datetime,date,timedelta

                    test_str = input("FECHA AGENDA DD-MM-YY ?: ")
 
                    # initializing format
                    format = "%d-%m-%y"
                    EndDate = datetime.now()+timedelta(days=2)
                    #res = datetime.now()

                    # using try-except to check for truth value
                    try:
                       # res = bool(datetime.strptime(test_str, format))
                        res = datetime.strptime(test_str, format)
                        v_abuscar = test_str + str(iSala) + iTurno 
                        #break
                        if (res >= EndDate) and not f_EncontroCita(v_abuscar,agenda):
                          break
                        else:
                              if not (res >= EndDate):
                               print("Fecha debe ser con 2 dias de Anticipacion")
                               continue
                              else:
                                print("Se encontro una reservacion para ese dia, favor de introducir otra")
                                continue
                    except ValueError:
                      #res = False
                      print("Fecha Invalida ",test_str )
                      continue
                ########## Validacion FECHA #########


                #Se genera consecutivo de agenda
                #8 Car fecha + 1 Car Sala +  1 Car Turno + 1 Car cliente
                AgendaFolio = test_str + str(iSala) + iTurno + str(iCliente) 
                print("Folio de agenda = ",AgendaFolio)

                AgendaEventoNombre = input("Nombre Evento ?: ")
                agenda[AgendaFolio] = AgendaEventoNombre

                #manejo de SQL Lite 3 
                #20221031
                valores = (AgendaFolio, AgendaEventoNombre)
                sqlite_query = "INSERT INTO t_agenda VALUES(?,?)"
                p_EjecutaComandoSQL(sqlite_query ,valores)   
                #manejo de SQL Lite 3 
                print("")#Espacio
#  ************* SUBMENU 1. [1]. Registrar una nueva reservación **************** 
                
                
#  ************* SUBMENU 1. [2]. Modificiar descripción de una reservación ****************                
                print("") #espacio
            elif (opcion=='2'):
                print('Modificiar descripción de una reservación')
                print("") #espacio

                v_folio = input("Capture el folio de la Agenda a Modificar")
                v_loencontre = False
                v_evento = ""
                v_eventomodificado = ""

                for clave in agenda:
                  if (len(clave) >= 10):
                   if (clave == str(v_folio)):
                      #print(clave,"      ",clave[9:10],"      ",agenda[clave])
                      v_loencontre = True
                      v_evento = agenda[clave]

                if v_loencontre:
                  #v_eventomodificado = input("Para el Folio ",v_folio, "con _Descripcion " , v_evento , " Dame la nueva descripcion ?")
                  v_eventomodificado = input(" Dame la nueva descripcion ")
                  agenda[clave] = v_eventomodificado

                  #manejo de SQL Lite 3 
                  #20221031
                  valores = (v_eventomodificado, v_folio)
                  sqlite_query = """Update t_agenda set AgeNom = ? where Agecod = ?"""
                  p_EjecutaComandoSQL(sqlite_query ,valores)          
                  #manejo de SQL Lite 3 


                  print("La modificacion se realizo correctamente")
                else:
                  print("No existe el folio ", v_folio)
#  ************* SUBMENU 1. [2]. Modificiar descripción de una reservación **************** 


#  ************* SUBMENU 1. [3]. Consultar disponibilidad de salas para una fecha **************** 
            elif (opcion=='3'):
                print('Consultar disponibilidad de salas para una fecha')
                print("") #espacio
                v_fecha  = input("Dame la fecha ?")

                ##Recorer la Sala
                for clave_sala in sala:
                  v_encontre_M = False
                  v_encontre_T = False
                  v_encontre_N = False

                  for clave_agenda in agenda:
                    if (len(clave_agenda) >= 10):
                        #Para la Sala del cilco se revisa la fecha , la Sala y el Turno M
                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "M"  ):                 
                          v_encontre_M = True

                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "T"  ):                 
                          v_encontre_T = True

                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "N"  ):                 
                          v_encontre_N = True

                  if not (v_encontre_M):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno M")
                      
                  if not (v_encontre_T):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno T")

                  if not (v_encontre_N):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno N")
#  ************* SUBMENU 1. [3]. Consultar disponibilidad de salas para una fecha **************** 


#  ************* SUBMENU 1. [4]. Eliminar una reservación **************** 
            elif (opcion=='4'):
                print('Borrado de una reservación')
                print("") #espacio

                v_loencontre = False
                v_evento = ""
                v_eventomodificado = ""
                v_folio = ""

                #Validacion Fecha A bORRAR
                format = "%d-%m-%y"
                EndDate = datetime.now()+timedelta(days=3)
                #while True:
                v_folio = input("Capture el folio de la Agenda a Borrar")
                try:
                   v_solo_fecha =v_folio[0:8]
                   res = datetime.strptime(v_solo_fecha, format)
                   if (res >= EndDate) :
                     print("")
                     #break
                   else:
                     print("Fecha no es Mayor  Fecha actual + 3 dias")
                     #continue
                except ValueError:
                   #res = False
                   print("Fecha Invalida ",v_solo_fecha )
                   #print("")
                   #continue
                #Validacion Fecha A bORRAR


                for clave in agenda:
                  if (len(clave) >= 10):
                   if (clave == str(v_folio)):
                      #print(clave,"      ",clave[9:10],"      ",agenda[clave])
                      v_loencontre = True
                      v_evento = agenda[clave]

                if v_loencontre:
                    print(agenda[clave])
                    del agenda[v_folio]                

                    try:
                      sqliteConnection = sqlite3.connect('EvidenciaSQLlite.db')
                      cursor = sqliteConnection.cursor()
                      #print("Connected to SQLite")

                      sql_update_query = """DELETE from t_agenda where AgeCod = ?"""
                      cursor.execute(sql_update_query, (v_folio,))
                      sqliteConnection.commit()
                      print("Record deleted successfully")
                      cursor.close()
                    except sqlite3.Error as error:
                      print("Failed to delete reocord from a sqlite table", error)
                    finally:
                      if sqliteConnection:
                          sqliteConnection.close()

                    print("Se borro el Folio = ",v_folio)
                ########## Validacion FECHA #########
#  ************* SUBMENU 1. [4]. Eliminar una reservación **************** 


#  ************* SUBMENU 1. [5]. Saliendo del submenu de reservaciones **************** 
            elif (opcion=='5'):
                print(' ** Saliendo del submenu de reservaciones **')
                print("")#espacio
                opcion = '0'
                break
            else:
                print('No existe la opcion.')
        print("") #Espacio
#  ************* SUBMENU 1. [5]. Saliendo del submenu de reservaciones **************** 
#  ************* SUBMENU 1. RESERVACIONES ****************


#  ************* SUBMENU 2. REPORTES ****************
    if (opcion=='2'):
        opcion = '0' 

        while not(opcion=='9'):
            print('          *** SUBMENU DE REPORTES***             ')
            print(' [1]. Reporte en pantalla de reservaciones para una fecha')                 
            print(' [2]. Exportar reporte tabular en Excel') 
            print(' [3]. Salir') # en todas son 5 pero se veria mal?                                                
            print("") #espacio

            opcion=input('  --- ¿Cuál opcion?: ')
        
#  ************* SUBMENU 2. [1]. Reporte en pantalla de reservaciones para una fecha ****************         
            if (opcion=='1'):
                print('[1]. Reporte en pantalla de reservaciones para una fecha')
                print("") #espacio
    

                v_fecha=input('¿Que Fecha ?: ')

                print('*****************************************************')
                print(' **** Reporte de reservaciones el dia ',v_fecha,'****')
                print('Folio            Turno       Nombre Evento')
                print('*****************************************************')
                for clave in agenda:
                    if (len(clave) >= 10):
                      if (clave[0:8] == str(v_fecha)):
                        print(clave,"      ",clave[9:10],"      ",agenda[clave])
                print(' ****************** Fin del reporte *****************')
                print("") #Espacio
#  ************* SUBMENU 2. [1]. Reporte en pantalla de reservaciones para una fecha ****************               


#  ************* SUBMENU 2. [2]. Exportar reporte tabular en Excel **************** 
            elif (opcion=='2'):
                print('[2].Exportar reporte tabular en Excel')
                print("") #espacio
       
                libro = openpyxl.Workbook()
                hoja = libro["Sheet"] 

                hoja.title = "Reporte tabular"
                hoja['A1'].value='Reporte de reservaciones el dia '
                hoja['A2'].value=('Sala')
                #hoja['A3'].value= clave[8:9] #NombreSala
                hoja['B2'].value='Cliente'
                #hoja['B3'].value=clave[10:11] #NombreCliente
                hoja['C2'].value='Folio'
                #hoja['C3'].value=clave
                hoja['D2'].value='Evento'
                #hoja['D3'].value= agenda[clave] #AgendaEventoNombre
                hoja['E2'].value='Turno'
                #hoja['E3'].value=clave[9:10]

                v_renglon = 3
                for clave in agenda:
                    if (len(clave) >= 10):
                      #if (clave[0:8] == str(v_fecha)):
                          #print(clave,"      ",clave[9:10],"      ",agenda[clave])

                          #hoja.title = "Reporte tabular"
                          #hoja['A1'].value='Reporte de reservaciones el dia '
                          #hoja['A2'].value=('Sala')
                          cellref=hoja.cell(v_renglon, column=1)
                          cellref.value= clave[8:9] #NombreSala

                          #hoja['B2'].value='Cliente'
                          cellref=hoja.cell(v_renglon, column=2)
                          cellref.value=clave[10:11] #NombreCliente
                          #hoja['C2'].value='Folio'

                          cellref=hoja.cell(v_renglon, column=3)
                          cellref.value=clave
                          #hoja['D2'].value='Evento'

                          cellref=hoja.cell(v_renglon, column=4)
                          cellref.value= agenda[clave] #AgendaEventoNombre
                          #hoja['E2'].value='Turno'

                          cellref=hoja.cell(v_renglon, column=5)
                          cellref.value=clave[9:10]

                          v_renglon = v_renglon + 1

                libro.save('ReporteTabular.xlsx')
                print('Libro creado exitosamente!')
#  ************* SUBMENU 2. [2]. Exportar reporte tabular en Excel **************** 


#  ************* SUBMENU 2. [3]. Saliendo del submenu de reportes **************** 
            elif (opcion=='3'):
                print(' ** Saliendo del submenu de reportes  **')
                print("") #Espacio
                opcion = '0'
                break
            else:
                print('No existe la opcion..')
                print("") #Espacio
#  ************* SUBMENU 2. [3]. Saliendo del submenu de reportes ****************
#  ************* [1]. RESERVACIONES ****************


#  ************* [4]. REGISTAR UN CLIENTES ****************
    elif (opcion=='4'):
        print(' **** Cliente ****')
        #print(cliente) Andy 15-10-22 
        print("") #Espacio

        #manejo de SQL Lite 3 
        #20221029
        siguiente = f_SiguienteValorContador('Clientes')
        print("Debug siguiente llamado a l a funcion,")
        print(siguiente)

        #manejo de SQL Lite 3 

        siguiente =  siguiente + 1
        print("Debug siguiente sumado")
        print(siguiente)


        if (siguiente == 1):
          #Inserta el Nuevo contador
          #manejo de SQL Lite 3 
          #20221031
          valores = ('Clientes', siguiente)
          sqlite_query = "INSERT INTO t_contador VALUES(?,?)"
          p_EjecutaComandoSQL(sqlite_query ,valores)   
          #manejo de SQL Lite 3 
    

        else:
          ##Actualiza el contador
          # update value
          #siguiente =  siguiente + 1
          #contador['Clientes'] = siguiente

          #manejo de SQL Lite 3 
          #20221031
          valores = (siguiente,  'Clientes')
          sqlite_query = """Update t_contador set ConUlt = ? where ConCod = ?"""
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 

    
        print('Nuevo Cliente',siguiente)

        NombreCliente = ""
        while f_ValidaCadenaVacia(NombreCliente):
          NombreCliente=input('¿Nombre cliente?: ')
          if f_ValidaCadenaVacia(NombreCliente):
            print("Nose puede dejar vacio, ingresa un nombre")
 

        #sala['SalCod'] = siguiente
        #sala['SalNam'] = NombreCliente

        #sigue actualizando la estructura
        cliente[siguiente] = NombreCliente

        #manejo de SQL Lite 3 
        #20221031
        valores = (siguiente,NombreCliente)
        sqlite_query = "INSERT INTO t_cliente VALUES(?,?)"
        p_EjecutaComandoSQL(sqlite_query ,valores)          
        #manejo de SQL Lite 3         

        #print(siguiente)
        print("Cliente registrado correctamente") #Andy 15-10-22 Quitare lo que muestralas cadenas de cliente para que solo se vea que se ingreso correctamente
        print("") #Andy 15-10-22 Para que no se vea todo junto
#  ************* [4]. REGISTAR UN CLIENTES ****************


#  ************* [3]. REGISTAR UNA SALA ****************
    elif (opcion=='3'):
        print(' **** Salas ****')
        print("") #Espacio
        #print(sala) Andy 15-10-22 Para que no se vean las listas al usuario
        
        #manejo de SQL Lite 3 
        #20221029
        siguiente = f_SiguienteValorContador('Salas')
        print("Debug siguiente llamado a l a funcion,")
        print(siguiente)

        #manejo de SQL Lite 3 

        siguiente =  siguiente + 1
        print("Debug siguiente sumado")
        print(siguiente)


        if (siguiente == 1):
          #Inserta el Nuevo contador
          #manejo de SQL Lite 3 
          #20221031
          valores = ('Salas', siguiente)
          sqlite_query = "INSERT INTO t_contador VALUES(?,?)"
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 


        else:
          ##Actualiza el contador
          # update value
          #siguiente =  siguiente + 1
          #contador['Clientes'] = siguiente

          #manejo de SQL Lite 3 
          #20221031
          valores = (siguiente,  'Salas')
          sqlite_query = """Update t_contador set ConUlt = ? where ConCod = ?"""
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 
  


        print('Nueva Sala',siguiente)
        

        NombreSala = ""
        while f_ValidaCadenaVacia(NombreSala):
           NombreSala=input('¿Nombre Sala?: ')
           if f_ValidaCadenaVacia(NombreSala):
             print("No se puede dejar vacio, ingresa un nombre a la sala")
             NombreSala=input('¿Nombre sala?: ')

        CapacidadSala=0


        #####Validacion de Capacidad
        while True:
          try:
             CapacidadSala=int(input(f'¿Capacidad Sala?: '))
             if  (CapacidadSala <= 0):
               print(f"Debes escribir un número mayor a 0")
               #CapacidadSala=int(input('¿Capacidad Sala?: '))  
             else:
               break
          except ValueError:
            print("Favor de ingresar un dato valido")
            continue
        #####Validacion de Capacidad  


        #sala['SalCod'] = siguiente
        #sala['SalNam'] = NombreCliente

        #Sigue Actualizando la Sala
        sala[siguiente] = NombreSala

        #salaCapacidad[siguiente] = CapacidadSala

        #manejo de SQL Lite 3 
        #20221031
        valores = (siguiente,NombreSala,CapacidadSala)
        sqlite_query = "INSERT INTO t_sala VALUES(?,?,?)"
        p_EjecutaComandoSQL(sqlite_query ,valores)          
        #manejo de SQL Lite 3 

        #print(siguiente)
        print("Sala creada correctamente") #Andy 15-10-22 
        #print(salaCapacidad) Andy 15-10-22 Esto era lo que mostraba las llaves pero no se ocupan para el usuario
        
        # Andy 16-10-22 La capacidad de la sala debe ser mayor a 0
        print("")#Espacio
#  ************* [3]. REGISTAR UNA SALA ****************


#  ************* [5]. SALIENDO DEL MENU ****************
    elif (opcion=='5'):
        print(' ** Saliendo del menu  **')
    #else:
        #print('No existe la opcion...')